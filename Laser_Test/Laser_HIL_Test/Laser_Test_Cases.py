###################################################################################
# Author  : Sabry Elsayed
#  
# Data    : 25 /3 / 2024
# Version : V.0.0
####################################################################################

################################## This is Laser Range Finder Hardware in the loop Test ####################################
import Laser
import openpyxl 
from Laser import *

from contextlib import ContextDecorator

########################################## To Trace Functions ##########################################
class TraceBack(ContextDecorator):
    def __init__(self, name):
        self.name = name 

    def __enter__(self):
        print("---------------------------------------","Enter", self.name,"-----------------------------------------------------------------")
        # print("Enter", self.name)
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        print("---------------------------------------","Exit", self.name,"-----------------------------------------------------------------")
        # print("Exit", self.name)
#############################################################################################################              
      
###################################################### Each State Response ###################################
expected_response = [0x1, 0x0, 0x0, 0x4, 0x1]
expected_Err_response = [0x00, 0x55]
expexted_Not_Data_Frame_Err = [0x00, 0x66]
expexted_Wrong_Response = [0x00, 0x77]
###############################################################################################################

###################################################### Open Excel Sheet #######################################
# Open the Excel workbook
path = 'D:\\Laser_Hil_Test\\Laser_HIL_Test_4\\Laser_Test_Report.xlsx'
workbook = openpyxl.load_workbook(path)
################################################################################################################
# Select the worksheet
sheet = workbook['Sheet1']  # Replace 'Sheet1' with the name of your sheet
first_row_index = 2  # writing to row 
#################################################################################################################

##################################################### Begin Test Case 1 ##########################################
@TraceBack("test_1")
def test_1():
    global first_row_index
    #Clear Buffers and receive command 
    can_serial.flushInput()
    can_serial.flushOutput()
    LR_Process_CMD()     
    LR_Process_CMD()
    can_serial.flushInput()
    can_serial.flushOutput()
    
    #Send Data Frame 
    WriteDataToSerialPort(bytearray(LR_CONTINOUS_MODE_RESPONSE))
    WriteDataToSerialPort(bytearray(LR_CONTINOUS_MODE_RESPONSE))

    #Receive Data and check on expected and actual result 
    response = LR_ReceiveResponse()   
    response = LR_ReceiveResponse() 

    if response is not None:
        response = [int(value, 16) for value in response]
        if response != None:
            print("Received Response:" , response)
            print("Expected Response:",expected_response)
            first_row_index+=1
            sheet.cell(row=2, column=5).value = str(response)  # Write actual data to column E
            retval = check_response(response[1:-2] , expected_response)
            if retval == True:
                        sheet.cell(row=2, column=6).value = "Pass"  # Write actual data to column E
                        print("Laser Test case 1 : pass")
            else:
                        sheet.cell(row=2, column=6).value = "Fail"  # Write actual data to column E
                        print("Laser Test case 1 : fail")
            workbook.save(path)            
##################################################### End Case 1 ###############################################################
                        
##################################################### Begin Test Case 2 ###############################################################
@TraceBack("test_2")
def test_2():
        global first_row_index
        #Clear Buffers and sleep until 10 second to indicate that communication Error State 
        can_serial.flushInput()
        time.sleep(10)
        i = 0
        can_serial.flushInput()

        #Receive Data and check on expected and actual result
        while  i < 2 :
          response = LR_ReceiveResponse()
          i+=1

        if response is not None:     
            response = [int(value, 16) for value in response]
            if response != None:
                print("Received Response:" , response)
                print("Expected Response:",expected_Err_response)
                first_row_index+=1
                sheet.cell(row=3, column=5).value = str(response)  # Write actual data to column E
                retval = check_response(response[1:-5] , expected_Err_response)
                if retval == True:
                            sheet.cell(row=3, column=6).value = "Pass"  # Write actual data to column E
                            print("Laser Test case 2 : pass") 
                else:
                    with open('test_results.txt', 'a') as file:
                            sheet.cell(row=3, column=6).value = "Fail"  # Write actual data to column E
                            print("Laser Test case 2 : fail")
                workbook.save(path)            
##################################################### End Test Case 2 ###############################################################


##################################################### Begin Test Case 3 ###############################################################
@TraceBack("test_3")     
def test_3():
        global first_row_index
        #Clear Buffers and receive command 
        can_serial.flushInput()
        can_serial.flushOutput()
        LR_Process_CMD()     
        LR_Process_CMD()
        can_serial.flushInput()
        can_serial.flushOutput()
        
        #Receive Data and check on expected and actual result
        i = 0
        while  i < 40 :
            WriteDataToSerialPort(bytearray(LR_NOT_DATA_FRAME_CMD))
            response = LR_ReceiveResponse()
            time.sleep(.3)
            # print("data was sent")
            i+=1
        can_serial.flushInput()  
        response = LR_ReceiveResponse()

        if response is not None:
            response = [int(value, 16) for value in response]
            if response != None:
                print("Received Response:" , response)
                print("Expected Response:",expexted_Not_Data_Frame_Err)
                first_row_index+=1
                sheet.cell(row=4, column=5).value = str(response)  # Write actual data to column E
                retval = check_response(response[1:-5] , expexted_Not_Data_Frame_Err)
                if retval == True:
                            sheet.cell(row=4, column=6).value = "Pass"  # Write actual data to column E
                            print("Laser Test case 3 : pass")
                else:
                            sheet.cell(row=4, column=6).value = "Fail"  # Write actual data to column E
                            print("Laser Test case 3 : fail")
                workbook.save(path)            
##################################################### End Test Case 3 ###############################################################
                            

##################################################### Begin Test Case 4 ###############################################################                            
@TraceBack("test_4")
def test_4():
        global first_row_index
        #Clear Buffers and receive command 
        can_serial.flushInput()
        can_serial.flushOutput()
      

        #Receive Data and check on expected and actual result
        i = 0
        while  i < 40 :
            WriteDataToSerialPort(bytearray(LR_WRONG_RESPONSE_FRAME))
            response = LR_ReceiveResponse()
            time.sleep(.3)
            # print("data was sent")
            i+=1
        can_serial.flushInput()  
        response = LR_ReceiveResponse()  

        if response is not None:
            response = [int(value, 16) for value in response]
            if response != None:
                print("Received Response:" , response)
                print("Expected Response:",expexted_Wrong_Response)
                first_row_index+=1
                sheet.cell(row=5, column=5).value = str(response)  # Write actual data to column E
                retval = check_response(response[1:-5] , expexted_Wrong_Response)
                if retval == True:
                            sheet.cell(row=5, column=6).value = "Pass"  # Write actual data to column E
                            print("Laser Test case 4 : pass") 
                else:
                            sheet.cell(row=5, column=6).value = "Fail"  # Write actual data to column E
                            print("Laser Test case 4 : fail")
                workbook.save(path)            
##################################################### End Test Case 4 ###############################################################

def main():
    test_1()
    test_2()
    test_3()
    test_4()
  


# if __name__ == "Laser_Test_Cases":
main()




















